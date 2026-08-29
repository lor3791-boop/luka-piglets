# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---------- Sheet 1: data entry ----------
ws = wb.active
ws.title = "전입전출입력"

HEADERS = ["일자", "구분", "세부구분", "농장명", "두수"]
HEADER_FILL = PatternFill("solid", fgColor="2C4A40")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
EXAMPLE_FONT = Font(name="Arial", size=10, italic=True, color="808080")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
IN_FARM_NAME = "익산 자돈사"

for col, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER

widths = [14, 10, 12, 20, 10]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:E1"

# example rows (clearly marked; delete before real use)
examples = [
    ["2026-09-01", "전출", "농장전출", "풍년농장", 600],
    ["2026-09-03", "전입", "", IN_FARM_NAME, 700],
    ["2026-09-05", "전출", "외부판매", "미래축산", 84],
]
for r, row in enumerate(examples, start=2):
    for col, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = EXAMPLE_FONT
        cell.border = BORDER
        if col == 1:
            cell.number_format = "@"  # keep date as text, avoid Excel auto date-serial

# Pre-format date & farm-name columns as text for the next many rows so values
# typed in stay literal text (avoids Excel silently converting 2026-09-01 to a
# date serial, which the importer would otherwise have to special-case).
for r in range(2, 400):
    ws.cell(row=r, column=1).number_format = "@"
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = BORDER
        if r > len(examples) + 1:
            ws.cell(row=r, column=col).font = BODY_FONT

# data validation dropdowns
dv_type = DataValidation(type="list", formula1='"전입,전출"', allow_blank=False, showErrorMessage=True,
                          errorTitle="입력 오류", error="전입 또는 전출만 입력할 수 있습니다.")
dv_type.add("B2:B400")
ws.add_data_validation(dv_type)

dv_sub = DataValidation(type="list", formula1='"농장전출,외부판매"', allow_blank=True, showErrorMessage=True,
                         errorTitle="입력 오류", error="농장전출 또는 외부판매만 입력할 수 있습니다. 전입이면 비워두세요.")
dv_sub.add("C2:C400")
ws.add_data_validation(dv_sub)

dv_head = DataValidation(type="whole", operator="greaterThan", formula1=0, allow_blank=True,
                          showErrorMessage=True, errorTitle="입력 오류", error="두수는 1 이상의 숫자여야 합니다.")
dv_head.add("E2:E400")
ws.add_data_validation(dv_head)

# ---------- Sheet 2: instructions ----------
ws2 = wb.create_sheet("작성방법")
ws2.column_dimensions["A"].width = 90
title = ws2.cell(row=1, column=1, value="전입전출입력 시트 작성 안내")
title.font = Font(name="Arial", size=13, bold=True, color="2C4A40")

lines = [
    "",
    "1. 이 파일은 '루카 후기자돈사 전입출 현황판' 웹사이트에 그대로 불러올 수 있는 고정 양식입니다.",
    "   시트 이름, 열 순서, 헤더 문구를 바꾸면 자동 인식이 되지 않으니 그대로 유지해 주세요.",
    "",
    "2. 열 설명 (전입전출입력 시트)",
    "   - 일자: YYYY-MM-DD 형식 (예: 2026-09-01). 열 서식이 '텍스트'로 미리 지정되어 있습니다.",
    "   - 구분: '전입' 또는 '전출' 중 선택 (드롭다운 제공).",
    "   - 세부구분: 구분이 '전출'일 때만 '농장전출' 또는 '외부판매' 선택. '전입'이면 비워둡니다.",
    "   - 농장명: 전출 시 실제 보낸 농장/거래처 이름을 입력합니다. 전입은 항상 '익산 자돈사'에서 들어오므로",
    "     비워두어도 웹사이트에서 자동으로 '익산 자돈사'로 처리됩니다.",
    "   - 두수: 1 이상의 숫자만 입력합니다.",
    "",
    "3. 예시로 들어있는 2~4행은 실제 데이터가 아닙니다. 입력을 시작하기 전에 지우거나 덮어써 주세요.",
    "",
    "4. 작성이 끝나면 파일을 저장한 뒤, 웹사이트의 '입력 · 현황' 탭에서 '엑셀 불러오기' 버튼으로",
    "   이 파일을 선택하세요. 불러오기를 누르면 현재 웹사이트에 있던 전입·전출 내역이 이 파일의",
    "   내용으로 전부 교체되니, 웹사이트에서만 입력해 둔 최근 내역이 있다면 먼저 이 파일에도",
    "   옮겨 적은 뒤 불러오세요.",
    "",
    "5. .xlsx 형식 그대로 올리면 됩니다 (다른 이름으로 저장할 필요 없음).",
]
for i, text in enumerate(lines, start=2):
    cell = ws2.cell(row=i, column=1, value=text)
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(wrap_text=False, vertical="top")

wb.save("output.xlsx")
print("saved")
